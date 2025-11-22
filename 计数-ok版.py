import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def process_excel_data(file_path):
    """
    直接处理Excel文件，按照指定尺寸顺序排序，并自动调整列宽
    """
    # 读取Excel文件
    df = pd.read_excel(file_path)

    # 从规格名称中提取颜色和尺寸
    def extract_color_size(name):
        # 提取不包含数字和字母的部分作为颜色
        color_match = re.match(r'[^A-Za-z0-9]*', str(name))
        if color_match:
            color = color_match.group(0).strip()  # 去除前后空格
            size = name[len(color):].strip()  # 去除前后空格
            return color, size
        return '', name.strip()  # 返回空字符串作为颜色，原始名称作为尺寸

    # 分离颜色和尺寸
    df[['颜色', '尺寸']] = df['规格名称'].apply(
        lambda x: pd.Series(extract_color_size(x))
    )

    # 分组汇总
    grouped = df.groupby(['规格编码', '颜色', '尺寸'])['规格数量'].sum().reset_index()
    grouped['尺寸数量'] = grouped['尺寸'] + '*' + grouped['规格数量'].astype(str)

    # 定义尺寸顺序
    size_order = ['S', 'M', 'L', 'XL', '2XL', '3XL', '4XL', '5XL', '6XL']

    # 按照指定尺寸顺序排序
    def sort_sizes(size_quantity_list):
        # 将尺寸数量字符串拆分为列表
        items = [item for item in size_quantity_list if item]

        # 创建排序键：按照size_order中的索引排序，不在顺序中的放在最后
        def get_sort_key(item):
            size = item.split('*')[0]  # 提取尺寸部分
            if size in size_order:
                return size_order.index(size)
            else:
                return len(size_order)  # 不在顺序中的放在最后

        # 按照尺寸顺序排序
        sorted_items = sorted(items, key=get_sort_key)

        # 使用中文逗号连接，最后一个使用中文逗号
        if len(sorted_items) > 1:
            # 前面部分用英文逗号，最后一个用中文逗号连接
            result = ','.join(sorted_items[:-1]) + '，' + sorted_items[-1]
        else:
            result = ''.join(sorted_items)

        return result

    # 生成最终结果
    result_df = grouped.groupby(['规格编码', '颜色'])['尺寸数量'].apply(
        lambda x: sort_sizes(x.tolist())
    ).reset_index()

    result_df['结果'] = result_df['规格编码'] + '-' + result_df['颜色'] + ' ：' + result_df['尺寸数量']

    return result_df


def auto_adjust_column_width(file_path):
    """
    自动调整Excel列宽
    """
    # 加载工作簿
    workbook = load_workbook(file_path)
    worksheet = workbook.active

    # 遍历所有列，自动调整列宽
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        # 计算每列的最大宽度
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        # 设置列宽，留出一些边距
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # 保存修改
    workbook.save(file_path)


# 使用示例
file_path = '/Users/lin/Documents/备货单2025-11-22.xlsx'
result = process_excel_data(file_path)
print(result)

# 保存结果到Excel
output_file = '汇总结果.xlsx'
result.to_excel(output_file, index=False)

# 自动调整列宽
auto_adjust_column_width(output_file)
print(f"结果已保存到 {output_file}，并已自动调整列宽")
