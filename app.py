from flask import Flask, request, jsonify, send_file
import pandas as pd
import re
import os
import tempfile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
import json
import sys

app = Flask(__name__)

# Vercel环境适配
if os.environ.get('VERCEL'):
    # 在Vercel上使用内存存储，但添加清理机制
    from collections import OrderedDict

    processed_data_store = OrderedDict()
    MAX_STORE_SIZE = 10  # 最多存储10个处理结果


    def cleanup_store():
        if len(processed_data_store) > MAX_STORE_SIZE:
            # 移除最旧的项目
            processed_data_store.popitem(last=False)
else:
    # 本地开发使用普通字典
    processed_data_store = {}


# 使用简单的内存存储（生产环境应使用数据库或会话）
# processed_data_store = {}  # 注释掉原来的定义
# 使用简单的内存存储（生产环境应使用数据库或会话）
# processed_data_store = {}
def process_excel_data(file_path):
    """
    直接处理Excel文件，按照指定尺寸顺序排序
    """
    # 读取Excel文件
    df = pd.read_excel(file_path)

    # 从规格名称中提取颜色和尺寸
    def extract_color_size(name):
        # 提取不包含数字和字母的部分作为颜色
        color_match = re.match(r'[^A-Za-z0-9]*', str(name))
        if color_match:
            color = color_match.group(0).strip()  # 去除前后空格
            size = name[len(color):].strip()  # 去除前后空格
            return color, size
        return '', name.strip()  # 返回空字符串作为颜色，原始名称作为尺寸

    # 分离颜色和尺寸
    df[['颜色', '尺寸']] = df['规格名称'].apply(
        lambda x: pd.Series(extract_color_size(x))
    )

    # 分组汇总
    grouped = df.groupby(['规格编码', '颜色', '尺寸'])['规格数量'].sum().reset_index()
    grouped['尺寸数量'] = grouped['尺寸'] + '*' + grouped['规格数量'].astype(str)

    # 定义尺寸顺序
    size_order = ['S', 'M', 'L', 'XL', '2XL', '3XL', '4XL', '5XL', '6XL']

    # 按照指定尺寸顺序排序
    def sort_sizes(size_quantity_list):
        # 将尺寸数量字符串拆分为列表
        items = [item for item in size_quantity_list if item]

        # 创建排序键：按照size_order中的索引排序，不在顺序中的放在最后
        def get_sort_key(item):
            size = item.split('*')[0]  # 提取尺寸部分
            if size in size_order:
                return size_order.index(size)
            else:
                return len(size_order)  # 不在顺序中的放在最后

        # 按照尺寸顺序排序
        sorted_items = sorted(items, key=get_sort_key)

        # 使用中文逗号连接，最后一个使用中文逗号
        if len(sorted_items) > 1:
            # 前面部分用英文逗号，最后一个用中文逗号连接
            result = ','.join(sorted_items[:-1]) + '，' + sorted_items[-1]
        else:
            result = ''.join(sorted_items)

        return result

    # 生成最终结果
    result_df = grouped.groupby(['规格编码', '颜色'])['尺寸数量'].apply(
        lambda x: sort_sizes(x.tolist())
    ).reset_index()

    result_df['结果'] = result_df['规格编码'] + '-' + result_df['颜色'] + ' ：' + result_df['尺寸数量']

    return result_df


def auto_adjust_column_width(file_path):
    """
    自动调整Excel列宽
    """
    # 加载工作簿
    workbook = load_workbook(file_path)
    worksheet = workbook.active

    # 遍历所有列，自动调整列宽
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        # 计算每列的最大宽度
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        # 设置列宽，留出一些边距
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # 保存修改
    workbook.save(file_path)


@app.route('/')
def index():
    """主页面"""
    # 获取当前日期用于默认文件名
    current_date = datetime.now().strftime('%Y-%m-%d')
    default_filename = f'备货单汇总{current_date}'

    return f'''
    <!DOCTYPE html>
    <html lang="zh-CN">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Excel数据处理与展示</title>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
        <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
        <style>
            body {{
                background-color: #f8f9fa;
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            }}
            .header {{
                background: linear-gradient(135deg, #0066cc 0%, #0099ff 100%);
                color: white;
                padding: 2rem 0;
                border-radius: 0 0 20px 20px;
                box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
            }}
            .card {{
                border-radius: 12px;
                border: none;
                box-shadow: 0 4px 12px rgba(0, 0, 0, 0.05);
                transition: transform 0.3s ease;
                margin-bottom: 20px;
            }}
            .card:hover {{
                transform: translateY(-5px);
            }}
            .upload-area {{
                border: 2px dashed #dee2e6;
                border-radius: 10px;
                padding: 2rem;
                text-align: center;
                cursor: pointer;
                transition: all 0.3s;
            }}
            .upload-area:hover, .upload-area.dragover {{
                border-color: #0066cc;
                background-color: #f0f8ff;
            }}
            .file-input {{
                display: none;
            }}
            .result-table {{
                width: 100%;
                border-collapse: collapse;
            }}
            .result-table th {{
                background-color: #e9ecef;
                position: sticky;
                top: 0;
            }}
            .table-container {{
                max-height: 400px;
                overflow-y: auto;
                border: 1px solid #dee2e6;
                border-radius: 8px;
            }}
            .btn-primary {{
                background: linear-gradient(135deg, #0066cc 0%, #0099ff 100%);
                border: none;
                border-radius: 8px;
                padding: 10px 20px;
            }}
            .btn-primary:hover {{
                background: linear-gradient(135deg, #0052a3 0%, #0077cc 100%);
            }}
            .loading {{
                display: none;
            }}
            .step {{
                display: flex;
                align-items: center;
                margin-bottom: 15px;
            }}
            .step-number {{
                width: 30px;
                height: 30px;
                border-radius: 50%;
                background-color: #0066cc;
                color: white;
                display: flex;
                align-items: center;
                justify-content: center;
                margin-right: 10px;
            }}
            .feature-icon {{
                width: 50px;
                height: 50px;
                background-color: #e6f2ff;
                border-radius: 50%;
                display: flex;
                align-items: center;
                justify-content: center;
                margin-right: 15px;
                color: #0066cc;
            }}
            .export-options {{
                background-color: #f8f9fa;
                border-radius: 8px;
                padding: 15px;
                margin-top: 15px;
            }}
        </style>
    </head>
    <body>
        <!-- 头部 -->
        <div class="header mb-5">
            <div class="container">
                <div class="row align-items-center">
                    <div class="col-md-8">
                        <h1><i class="fas fa-file-excel me-2"></i>Excel数据处理工具</h1>
                        <p class="lead">上传Excel文件，自动提取规格信息并生成汇总结果</p>
                    </div>
                    <div class="col-md-4 text-end">
                        <div class="feature-icon d-inline-flex">
                            <i class="fas fa-cogs fa-lg"></i>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <div class="container">
            <!-- 功能说明 -->
            <div class="row mb-5">
                <div class="col-md-4">
                    <div class="card h-100">
                        <div class="card-body">
                            <h5 class="card-title"><i class="fas fa-info-circle me-2"></i>功能说明</h5>
                            <div class="step">
                                <div class="step-number">1</div>
                                <span>上传包含规格信息的Excel文件</span>
                            </div>
                            <div class="step">
                                <div class="step-number">2</div>
                                <span>系统自动提取颜色和尺寸信息</span>
                            </div>
                            <div class="step">
                                <div class="step-number">3</div>
                                <span>按指定尺寸顺序排序并汇总</span>
                            </div>
                            <div class="step">
                                <div class="step-number">4</div>
                                <span>生成格式化结果并下载</span>
                            </div>
                        </div>
                    </div>
                </div>
                <div class="col-md-8">
                    <div class="card h-100">
                        <div class="card-body">
                            <h5 class="card-title"><i class="fas fa-upload me-2"></i>上传Excel文件</h5>
                            <div class="upload-area" id="uploadArea">
                                <i class="fas fa-cloud-upload-alt fa-3x text-muted mb-3"></i>
                                <h5>拖放文件到此处或点击上传</h5>
                                <p class="text-muted">支持 .xlsx 格式的Excel文件</p>
                                <button class="btn btn-primary mt-2">选择文件</button>
                                <input type="file" id="fileInput" class="file-input" accept=".xlsx">
                            </div>
                            <div class="loading mt-3 text-center" id="loading">
                                <div class="spinner-border text-primary" role="status">
                                    <span class="visually-hidden">处理中...</span>
                                </div>
                                <p class="mt-2">正在处理文件，请稍候...</p>
                            </div>
                        </div>
                    </div>
                </div>
            </div>

            <!-- 结果显示区域 -->
            <div class="row">
                <div class="col-12">
                    <div class="card">
                        <div class="card-body">
                            <h5 class="card-title"><i class="fas fa-table me-2"></i>处理结果</h5>
                            <div id="resultArea" class="table-container">
                                <p class="text-center text-muted p-4">上传Excel文件后，处理结果将在此显示</p>
                            </div>
                            <div class="mt-3 text-end" id="downloadArea" style="display: none;">
                                <div class="export-options">
                                    <div class="row align-items-center">
                                        <div class="col-md-6">
                                            <label for="fileName" class="form-label">文件名：</label>
                                            <input type="text" id="fileName" class="form-control" value="{default_filename}" placeholder="请输入文件名">
                                        </div>
                                        <div class="col-md-6 text-end">
                                            <button class="btn btn-success" id="downloadBtn">
                                                <i class="fas fa-download me-2"></i>下载Excel文件
                                            </button>
                                            <button class="btn btn-outline-primary ms-2" id="downloadCsvBtn">
                                                <i class="fas fa-file-csv me-2"></i>下载CSV
                                            </button>
                                        </div>
                                    </div>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <!-- 页脚 -->
        <footer class="mt-5 py-4 bg-light">
            <div class="container text-center">
                <p class="mb-0 text-muted">Excel数据处理工具 &copy; 2023</p>
            </div>
        </footer>

        <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
        <script>
            // 全局变量存储处理结果的ID
            let currentResultId = null;

            document.addEventListener('DOMContentLoaded', function() {{
                const uploadArea = document.getElementById('uploadArea');
                const fileInput = document.getElementById('fileInput');
                const resultArea = document.getElementById('resultArea');
                const loading = document.getElementById('loading');
                const downloadArea = document.getElementById('downloadArea');
                const downloadBtn = document.getElementById('downloadBtn');
                const downloadCsvBtn = document.getElementById('downloadCsvBtn');
                const fileNameInput = document.getElementById('fileName');

                // 点击上传区域触发文件选择
                uploadArea.addEventListener('click', function() {{
                    fileInput.click();
                }});

                // 文件选择变化
                fileInput.addEventListener('change', function(e) {{
                    if (e.target.files.length > 0) {{
                        processExcelFile(e.target.files[0]);
                    }}
                }});

                // 拖放功能
                uploadArea.addEventListener('dragover', function(e) {{
                    e.preventDefault();
                    uploadArea.classList.add('dragover');
                }});

                uploadArea.addEventListener('dragleave', function() {{
                    uploadArea.classList.remove('dragover');
                }});

                uploadArea.addEventListener('drop', function(e) {{
                    e.preventDefault();
                    uploadArea.classList.remove('dragover');
                    if (e.dataTransfer.files.length > 0) {{
                        processExcelFile(e.dataTransfer.files[0]);
                    }}
                }});

                // 处理Excel文件
                function processExcelFile(file) {{
                    if (!file.name.endsWith('.xlsx')) {{
                        alert('请上传.xlsx格式的Excel文件');
                        return;
                    }}

                    loading.style.display = 'block';
                    resultArea.innerHTML = '<p class="text-center text-muted p-4">正在处理文件，请稍候...</p>';
                    downloadArea.style.display = 'none';

                    const formData = new FormData();
                    formData.append('file', file);

                    fetch('/process', {{
                        method: 'POST',
                        body: formData
                    }})
                    .then(response => response.json())
                    .then(data => {{
                        if (data.success) {{
                            currentResultId = data.result_id; // 保存结果ID
                            displayResults(data.results);
                        }} else {{
                            resultArea.innerHTML = '<p class="text-center text-danger p-4">' + data.error + '</p>';
                        }}
                        loading.style.display = 'none';
                    }})
                    .catch(error => {{
                        console.error('处理文件时出错:', error);
                        loading.style.display = 'none';
                        resultArea.innerHTML = '<p class="text-center text-danger p-4">处理文件时出错，请检查文件格式是否正确</p>';
                    }});
                }}

                // 显示处理结果
                function displayResults(results) {{
                    if (results.length === 0) {{
                        resultArea.innerHTML = '<p class="text-center text-muted p-4">未找到可处理的数据</p>';
                        return;
                    }}

                    let tableHTML = `
                        <table class="table table-striped result-table">
                            <thead>
                                <tr>
                                    <th>序号</th>
                                    <th>规格编码</th>
                                    <th>颜色</th>
                                    <th>汇总结果</th>
                                </tr>
                            </thead>
                            <tbody>
                    `;

                    results.forEach((item, index) => {{
                        tableHTML += `
                            <tr>
                                <td>${{index + 1}}</td>
                                <td>${{item.规格编码}}</td>
                                <td>${{item.颜色}}</td>
                                <td>${{item.结果}}</td>
                            </tr>
                        `;
                    }});

                    tableHTML += `
                            </tbody>
                        </table>
                    `;

                    resultArea.innerHTML = tableHTML;
                    downloadArea.style.display = 'block';
                }}

                // 下载Excel文件
                downloadBtn.addEventListener('click', function() {{
                    if (!currentResultId) {{
                        alert('请先上传并处理Excel文件');
                        return;
                    }}

                    const fileName = fileNameInput.value || '备货单汇总';
                    window.open('/download/excel?result_id=' + currentResultId + '&filename=' + encodeURIComponent(fileName), '_blank');
                }});

                // 下载CSV文件
                downloadCsvBtn.addEventListener('click', function() {{
                    if (!currentResultId) {{
                        alert('请先上传并处理Excel文件');
                        return;
                    }}

                    const fileName = fileNameInput.value || '备货单汇总';
                    window.open('/download/csv?result_id=' + currentResultId + '&filename=' + encodeURIComponent(fileName), '_blank');
                }});
            }});
        </script>
    </body>
    </html>
    '''


@app.route('/process', methods=['POST'])
def process_excel():
    """处理上传的Excel文件"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '没有上传文件'})

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '没有选择文件'})

        if not file.filename.endswith('.xlsx'):
            return jsonify({'success': False, 'error': '请上传.xlsx格式的Excel文件'})

        # 保存临时文件
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, file.filename)
        file.save(temp_path)

        # 处理Excel文件
        result_df = process_excel_data(temp_path)

        # 生成唯一ID用于存储处理结果
        import uuid
        result_id = str(uuid.uuid4())

        # 将处理结果存储到内存中
        processed_data_store[result_id] = {
            'dataframe': result_df.to_dict('records'),
            'columns': list(result_df.columns)
        }

        # 删除临时文件
        os.remove(temp_path)

        # 将结果转换为字典列表
        results = result_df.to_dict('records')

        return jsonify({
            'success': True,
            'results': results,
            'result_id': result_id
        })

    except Exception as e:
        return jsonify({'success': False, 'error': f'处理文件时出错: {str(e)}'})


@app.route('/download/excel')
def download_excel():
    """下载Excel格式的结果文件"""
    try:
        result_id = request.args.get('result_id')
        filename = request.args.get('filename', '备货单汇总')

        if not result_id or result_id not in processed_data_store:
            return jsonify({'success': False, 'error': '未找到处理结果，请先上传并处理文件'})

        # 从存储中获取处理结果
        result_data = processed_data_store[result_id]
        result_df = pd.DataFrame(result_data['dataframe'])

        # 创建Excel文件
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            result_df.to_excel(writer, index=False, sheet_name='汇总结果')

            # 获取工作表并调整列宽
            worksheet = writer.sheets['汇总结果']
            for idx, col in enumerate(result_df.columns):
                max_length = max(result_df[col].astype(str).str.len().max(), len(col)) + 2
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_length

        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f'{filename}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'success': False, 'error': f'下载文件时出错: {str(e)}'})


@app.route('/download/csv')
def download_csv():
    """下载CSV格式的结果文件"""
    try:
        result_id = request.args.get('result_id')
        filename = request.args.get('filename', '备货单汇总')

        if not result_id or result_id not in processed_data_store:
            return jsonify({'success': False, 'error': '未找到处理结果，请先上传并处理文件'})

        # 从存储中获取处理结果
        result_data = processed_data_store[result_id]
        result_df = pd.DataFrame(result_data['dataframe'])

        # 创建CSV文件
        output = BytesIO()
        result_df.to_csv(output, index=False, encoding='utf-8-sig')
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f'{filename}.csv',
            mimetype='text/csv'
        )

    except Exception as e:
        return jsonify({'success': False, 'error': f'下载文件时出错: {str(e)}'})


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=7100)

